import sqlite3
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

# =========================
# 設定
# =========================
DB_PATH = "genba.db"
EXCEL_PATH = "人工集計表.xlsx"
OUTPUT_PATH = "人工集計表_出力.xlsx"

# =========================
# UI用DB接続
# =========================
conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

# =========================
# マスタ取得
# =========================
cur.execute("SELECT worker_id, name FROM workers")
workers = cur.fetchall()

cur.execute("SELECT site_id, site_name FROM sites")
sites = cur.fetchall()

# =========================
# UI作成
# =========================
root = tk.Tk()
root.title("現場勤怠システム")

# -------------------------
# 名前
# -------------------------
tk.Label(root, text="名前").grid(row=0, column=0)

worker_var = tk.StringVar()
worker_box = ttk.Combobox(root, textvariable=worker_var)
worker_box["values"] = [f"{w[0]}:{w[1]}" for w in workers]
worker_box.grid(row=0, column=1)

# -------------------------
# 現場
# -------------------------
tk.Label(root, text="現場").grid(row=1, column=0)

site_var = tk.StringVar()
site_box = ttk.Combobox(root, textvariable=site_var)
site_box["values"] = [f"{s[0]}:{s[1]}" for s in sites]
site_box.grid(row=1, column=1)

# -------------------------
# 日付
# -------------------------
tk.Label(root, text="日付").grid(row=2, column=0)

date_entry = tk.Entry(root)
date_entry.grid(row=2, column=1)

# -------------------------
# type
# -------------------------
tk.Label(root, text="type(1人工/2残業/3深夜)").grid(row=3, column=0)

type_var = tk.StringVar(value="1")
type_box = ttk.Combobox(root, textvariable=type_var)
type_box["values"] = ["1", "2", "3"]
type_box.grid(row=3, column=1)

# =========================
# 出勤登録
# =========================
def insert_attendance():
    worker_id = int(worker_var.get().split(":")[0])
    site_id = int(site_var.get().split(":")[0])
    work_date = date_entry.get()
    type_ = int(type_var.get())

    cur.execute("""
        INSERT INTO attendance (worker_id, site_id, work_date, type, man_power)
        VALUES (?, ?, ?, ?, 1)
    """, (worker_id, site_id, work_date, type_))

    conn.commit()
    print("出勤登録完了")

# =========================
# Excel出力
# =========================
def export_excel():

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # ---------------------
    # 工事情報
    # ---------------------
    cur.execute("""
    SELECT project_name, year, month
    FROM project
    WHERE id = 1
    """)
    project_name, year, month = cur.fetchone()

    ws["A1"] = project_name
    ws["J1"] = f"{year}年 {month}月"

    # ---------------------
    # データ取得
    # ---------------------
    cur.execute("""
    SELECT worker_id, work_date, type, man_power
    FROM attendance
    """)
    rows = cur.fetchall()

    cur.execute("SELECT worker_id, name FROM workers")
    worker_map = dict(cur.fetchall())

    # ---------------------
    # 行・列自動取得
    # ---------------------
    name_row = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            name_row[v] = r

    date_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if v:
            date_col[str(v)] = c

    # ---------------------
    # 書き込み
    # ---------------------
    for worker_id, work_date, type_, man_power in rows:

        name = worker_map.get(worker_id)
        if not name:
            continue

        if name not in name_row:
            continue

        if work_date not in date_col:
            continue

        r = name_row[name]
        c = date_col[work_date]

        if type_ == 1:
            ws.cell(r, c).value = man_power

        elif type_ == 2:
            ws.cell(r + 10, c).value = man_power

        elif type_ == 3:
            ws.cell(r + 20, c).value = man_power

    wb.save(OUTPUT_PATH)
    print("Excel出力完了:", OUTPUT_PATH)

# =========================
# ボタン
# =========================
tk.Button(root, text="出勤登録", command=insert_attendance).grid(row=4, column=0)
tk.Button(root, text="Excel出力", command=export_excel).grid(row=4, column=1)

# =========================
# 起動
# =========================
root.mainloop()

conn.close()